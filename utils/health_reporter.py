"""
Reporting for the API Health Suite.

Responsibilities:
1. Store API health results.
2. Generate Excel report (with per-module breakdown).
3. Generate machine-readable summary.json (with per-module breakdown).

The JSON summary is intentionally compact and contains
the overall execution summary rather than the complete
request/response payloads.
"""

import json
import os
import re
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font as StylesFont, PatternFill, Alignment
except ImportError as e:
    raise ImportError(
        "openpyxl is required for health reporting. "
        "Add `openpyxl` to requirements.txt"
    ) from e

# Alias to the cell-styling Font so it is never shadowed by
# openpyxl.drawing.text.Font, which does NOT support `bold`.
Font = StylesFont


PASS_FILL = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid",
)

FAIL_FILL = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid",
)

SLA_WARN_FILL = PatternFill(
    start_color="FFEB9C",
    end_color="FFEB9C",
    fill_type="solid",
)

HEADER_FILL = PatternFill(
    start_color="4472C4",
    end_color="4472C4",
    fill_type="solid",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

MASKED_HEADER_KEYS = (
    "authorization",
    "cookie",
    "set-cookie",
)

_ILLEGAL_CHARACTERS_RE = re.compile(
    r"[\000-\010]|[\013-\014]|[\016-\037]"
)


def _sanitize_for_excel(value):
    """
    Remove control characters that openpyxl cannot write.
    """
    if value is None:
        return ""

    if not isinstance(value, str):
        return value

    return _ILLEGAL_CHARACTERS_RE.sub("", value)


def _mask_headers(headers: dict) -> dict:
    """
    Mask sensitive HTTP headers before storing them in reports.
    """
    masked = {}

    for key, value in (headers or {}).items():
        if key.lower() in MASKED_HEADER_KEYS:
            if isinstance(value, str):
                masked[key] = value[:15] + "...MASKED"
            else:
                masked[key] = "MASKED"
        else:
            masked[key] = value

    return masked


def _pretty(value) -> str:
    """
    Pretty-print dictionaries/lists/JSON strings.
    """
    if value is None:
        return ""

    if isinstance(value, (dict, list)):
        return json.dumps(value, indent=2)

    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return json.dumps(parsed, indent=2)
        except (ValueError, TypeError):
            return value

    return str(value)


class HealthReporter:

    def __init__(self, run_folder: str):
        self.run_folder = run_folder
        self.results = []

        # Make absolutely sure the run directory exists.
        os.makedirs(self.run_folder, exist_ok=True)

    # =========================================================
    # RECORD RESULT
    # =========================================================

    def record(
        self,
        name: str,
        action: str,
        method: str,
        path: str,
        params: dict | None,
        status_code: int | None,
        expected_status: int,
        passed: bool,
        elapsed_ms: float | None,
        sla_ms: int,
        api_message: str = "",
        request_headers: dict | None = None,
        request_payload: dict | None = None,
        response_body=None,
        candidate_name: str = "",
        candidate_email: str = "",
        error: str = "",
        critical: bool = False,
        module: str = "Uncategorized",
        submodule: str = "",
    ):
        sla_status = "N/A"

        if elapsed_ms is not None and sla_ms:
            sla_status = (
                "WITHIN SLA"
                if elapsed_ms <= sla_ms
                else "SLA BREACH"
            )

        self.results.append(
            {
                "name": name,
                "module": module or "Uncategorized",
                "submodule": submodule or "",
                "candidate_name": candidate_name or "N/A",
                "candidate_email": candidate_email or "N/A",
                "action": action,
                "method": method,
                "endpoint": path,
                "params": params or {},
                "status_code": status_code,
                "expected_status": expected_status,
                "passed": passed,
                "elapsed_ms": elapsed_ms,
                "sla_ms": sla_ms,
                "sla_status": sla_status,
                "api_message": api_message,
                "request_headers": _mask_headers(request_headers),
                "request_payload": request_payload or {},
                "response_body": response_body,
                "screenshot": "N/A (API-only check)",
                "error": error,
                "critical": critical,
                "timestamp": datetime.now().isoformat(
                    timespec="seconds"
                ),
            }
        )

    # =========================================================
    # SUMMARY
    # =========================================================

    @property
    def summary(self) -> dict:

        total = len(self.results)

        passed = sum(
            1
            for result in self.results
            if result["passed"]
        )

        failed = total - passed

        sla_breaches = [
            result["action"]
            for result in self.results
            if result["sla_status"] == "SLA BREACH"
        ]

        critical_failed = [
            result["action"]
            for result in self.results
            if result["critical"]
            and not result["passed"]
        ]

        failed_endpoints = [
            {
                "name": result["name"],
                "module": result["module"],
                "action": result["action"],
                "status_code": result["status_code"],
                "expected_status": result["expected_status"],
                "error": result["error"],
                "api_message": result["api_message"],
            }
            for result in self.results
            if not result["passed"]
        ]

        return {
            "total": total,
            "passed": passed,
            "failed": failed,
            "pass_rate": (
                round((passed / total) * 100, 1)
                if total
                else 0
            ),
            "critical_failed": critical_failed,
            "sla_breaches": sla_breaches,
            "failed_endpoints": failed_endpoints,
            "run_time": datetime.now().isoformat(
                timespec="seconds"
            ),
        }

    # =========================================================
    # MODULE BREAKDOWN
    # =========================================================

    @property
    def modules(self) -> dict:
        """
        Per-module breakdown: total/passed/failed/pass_rate/failed_apis.

        Used for:
          - the Excel "Module Summary" sheet
          - the email dashboard's module table
          - summary.json's "modules" section

        This is what lets you filter the Details sheet or slice the
        report by module and hand a clean, module-scoped view to the
        developer who owns that module.
        """
        grouped: dict[str, dict] = {}

        for result in self.results:
            mod = result.get("module") or "Uncategorized"

            grouped.setdefault(
                mod,
                {
                    "total": 0,
                    "passed": 0,
                    "failed": 0,
                    "failed_apis": [],
                },
            )

            grouped[mod]["total"] += 1

            if result["passed"]:
                grouped[mod]["passed"] += 1
            else:
                grouped[mod]["failed"] += 1
                grouped[mod]["failed_apis"].append(result["action"])

        for mod, stats in grouped.items():
            stats["pass_rate"] = (
                round((stats["passed"] / stats["total"]) * 100, 1)
                if stats["total"]
                else 0
            )

        return grouped

    # =========================================================
    # EXCEL REPORT
    # =========================================================

    def write_excel(self) -> str:

        os.makedirs(self.run_folder, exist_ok=True)

        wb = Workbook()

        # -----------------------------------------------------
        # SUMMARY SHEET
        # -----------------------------------------------------

        ws = wb.active
        ws.title = "Summary"

        summary = self.summary

        rows = [
            ("Run Time", summary["run_time"]),
            ("Total Endpoints", summary["total"]),
            ("Passed", summary["passed"]),
            ("Failed", summary["failed"]),
            ("Pass Rate", f"{summary['pass_rate']}%"),
            (
                "Critical Failures",
                ", ".join(summary["critical_failed"])
                or "None",
            ),
            (
                "SLA Breaches",
                ", ".join(summary["sla_breaches"])
                or "None",
            ),
        ]

        for index, (label, value) in enumerate(
            rows,
            start=1,
        ):
            ws.cell(
                row=index,
                column=1,
                value=label,
            ).font = Font(bold=True)

            ws.cell(
                row=index,
                column=2,
                value=_sanitize_for_excel(value),
            )

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 70

        # -----------------------------------------------------
        # MODULE SUMMARY SHEET
        # -----------------------------------------------------

        mod_ws = wb.create_sheet("Module Summary")

        mod_headers = [
            "Module",
            "Total",
            "Passed",
            "Failed",
            "Pass Rate",
            "Failed APIs",
        ]

        mod_ws.append(mod_headers)

        for column in range(1, len(mod_headers) + 1):
            cell = mod_ws.cell(row=1, column=column)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        mod_ws.freeze_panes = "A2"

        for module_name, stats in sorted(self.modules.items()):
            row_index = mod_ws.max_row + 1

            mod_ws.append(
                [
                    _sanitize_for_excel(module_name),
                    stats["total"],
                    stats["passed"],
                    stats["failed"],
                    f"{stats['pass_rate']}%",
                    _sanitize_for_excel(
                        ", ".join(stats["failed_apis"]) or "None"
                    ),
                ]
            )

            row_fill = PASS_FILL if stats["failed"] == 0 else FAIL_FILL

            for column in range(1, len(mod_headers) + 1):
                cell = mod_ws.cell(row=row_index, column=column)
                cell.fill = row_fill
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        mod_ws.column_dimensions["A"].width = 24
        mod_ws.column_dimensions["B"].width = 10
        mod_ws.column_dimensions["C"].width = 10
        mod_ws.column_dimensions["D"].width = 10
        mod_ws.column_dimensions["E"].width = 12
        mod_ws.column_dimensions["F"].width = 70

        # -----------------------------------------------------
        # DETAILS SHEET
        # -----------------------------------------------------

        details = wb.create_sheet("Details")

        headers = [
            "Module",
            "Submodule",
            "Candidate Name",
            "Candidate Email",
            "Action",
            "Method",
            "Endpoint",
            "API Status",
            "Expected Status",
            "Duration (ms)",
            "SLA (ms)",
            "SLA Status",
            "API Message",
            "Request Headers",
            "Request Payload",
            "Response Body",
            "Screenshot",
            "Error",
            "Timestamp",
        ]

        details.append(headers)

        for column in range(
            1,
            len(headers) + 1,
        ):
            cell = details.cell(
                row=1,
                column=column,
            )

            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        details.freeze_panes = "A2"

        for result in self.results:

            row_index = details.max_row + 1

            details.append(
                [
                    _sanitize_for_excel(result["module"]),
                    _sanitize_for_excel(result["submodule"]),
                    _sanitize_for_excel(
                        result["candidate_name"]
                    ),
                    _sanitize_for_excel(
                        result["candidate_email"]
                    ),
                    _sanitize_for_excel(
                        result["action"]
                    ),
                    _sanitize_for_excel(
                        result["method"]
                    ),
                    _sanitize_for_excel(
                        result["endpoint"]
                    ),
                    result["status_code"],
                    result["expected_status"],
                    result["elapsed_ms"],
                    result["sla_ms"],
                    _sanitize_for_excel(
                        result["sla_status"]
                    ),
                    _sanitize_for_excel(
                        result["api_message"]
                    ),
                    _sanitize_for_excel(
                        _pretty(
                            result["request_headers"]
                        )
                    ),
                    _sanitize_for_excel(
                        _pretty(
                            result["request_payload"]
                        )
                    ),
                    _sanitize_for_excel(
                        _pretty(
                            result["response_body"]
                        )
                    ),
                    _sanitize_for_excel(
                        result["screenshot"]
                    ),
                    _sanitize_for_excel(
                        result["error"]
                    ),
                    _sanitize_for_excel(
                        result["timestamp"]
                    ),
                ]
            )

            row_fill = (
                PASS_FILL
                if result["passed"]
                else FAIL_FILL
            )

            for column in range(
                1,
                len(headers) + 1,
            ):
                cell = details.cell(
                    row=row_index,
                    column=column,
                )

                cell.fill = row_fill
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            if result["sla_status"] == "SLA BREACH":
                # SLA Status is now column 12 (Module + Submodule added
                # two columns before it — was column 10 previously).
                details.cell(
                    row=row_index,
                    column=12,
                ).fill = SLA_WARN_FILL

        widths = {
            "A": 20,   # Module
            "B": 22,   # Submodule
            "C": 18,   # Candidate Name
            "D": 24,   # Candidate Email
            "E": 26,   # Action
            "F": 8,    # Method
            "G": 32,   # Endpoint
            "H": 11,   # API Status
            "I": 13,   # Expected Status
            "J": 13,   # Duration (ms)
            "K": 10,   # SLA (ms)
            "L": 13,   # SLA Status
            "M": 30,   # API Message
            "N": 34,   # Request Headers
            "O": 34,   # Request Payload
            "P": 40,   # Response Body
            "Q": 20,   # Screenshot
            "R": 40,   # Error
            "S": 20,   # Timestamp
        }

        for column_letter, width in widths.items():
            details.column_dimensions[
                column_letter
            ].width = width

        output_path = os.path.join(
            self.run_folder,
            "api_health_report.xlsx",
        )

        wb.save(output_path)

        print(
            f"[health-suite] Excel report created: "
            f"{output_path}"
        )

        return output_path

    # =========================================================
    # JSON SUMMARY
    # =========================================================

    def write_summary_json(self) -> str:
        """
        Writes the machine-readable health summary, including the
        per-module breakdown, as a real monitoring contract:

        {
          "total": ..., "passed": ..., "failed": ..., "pass_rate": ...,
          "critical_failed": [...], "sla_breaches": [...],
          "failed_endpoints": [...], "run_time": "...",
          "modules": {
            "Payroll": {"total": .., "passed": .., "failed": ..,
                        "pass_rate": .., "failed_apis": [...]},
            ...
          }
        }

        This MUST remain inside the HealthReporter class.
        """

        os.makedirs(self.run_folder, exist_ok=True)

        output_path = os.path.join(
            self.run_folder,
            "summary.json",
        )

        payload = dict(self.summary)
        payload["modules"] = self.modules

        with open(
            output_path,
            "w",
            encoding="utf-8",
        ) as file:

            json.dump(
                payload,
                file,
                indent=2,
                ensure_ascii=False,
            )

        print(
            f"[health-suite] JSON summary created: "
            f"{output_path}"
        )

        return output_path