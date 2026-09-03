from openpyxl.styles import Alignment , Font
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import json
import time

from utils.run_manager import (
    get_run_folder
)

RUN_FOLDER = get_run_folder()

EXCEL_FOLDER = os.path.join(
    RUN_FOLDER,
    "excel"
)

os.makedirs(
    EXCEL_FOLDER,
    exist_ok=True
)

REPORT_FILE = os.path.join(
    EXCEL_FOLDER,
    "test_execution_report.xlsx"
)

FIELD_LOG_FILE = os.path.join(
    EXCEL_FOLDER,
    "onboarding_data_log.xlsx"
)

SQLI_SHEET_NAME = "SQL Injection Report"

    
def create_excel_report():

    if not os.path.exists(REPORT_FILE):

        wb = Workbook()

        ws = wb.active

        ws.title = "Execution Report"

        ws.append([
            "Execution Time",
            "Run ID",
            "Environment",
            "Username",
            "Test Name",
            "Status",

            # Test-level context
            "Candidate Name",
            "Candidate Email",
            "Action",

            # API call details
            "Method",
            "Endpoint",
            "API Status",
            "Expected Status",
            "Duration(ms)",
            "SLA(ms)",
            "SLA Status",
            "API Message",
            "Request Headers",
            "Request Payload",
            "Response Body",

            # Failure details
            "Screenshot",
            "Error",
        ])

        wb.save(REPORT_FILE)

        print(
            f"[EXCEL CREATED] {REPORT_FILE}"
        )


def write_result(
    test_name,
    status,
    error="",
    screenshot="",
    action="",
    candidate_name="",
    candidate_email="",
    api_message="",
    run_id="",
    environment="",
    username="",
    method="",
    endpoint="",
    api_status="",
    expected_status="",
    duration="",
    sla="",
    sla_status="",
    request_headers="",
    request_payload="",
    response_body="",

):

    create_excel_report()

    wb = load_workbook(REPORT_FILE)

    ws = wb["Execution Report"]

    ws.append([
        str(datetime.now()),
        str(run_id),
        str(environment),
        str(username),
        test_name,
        status,

        candidate_name,
        candidate_email,
        action,

        str(method),
        str(endpoint),
        str(api_status),
        str(expected_status),
        str(duration),
        str(sla),
        str(sla_status),
        api_message,
        str(request_headers),
        str(request_payload),
        str(response_body),

        error,
        screenshot,
    ])

    wb.save(REPORT_FILE)

    print(
        f"[REPORT UPDATED] {REPORT_FILE}"
    )


def write_api_log(*,
        test_name,
        method,
        endpoint,
        status_code,
        run_id=None,
        environment=None,
        username=None,
        expected_status=None,
        actual_status=None,
        duration=None,
        sla=None,
        sla_status=None,
        result=None,
        request_headers=None,
        request_payload=None,
        response_body=None,
        error=""
):
    """
    BACKWARD-COMPATIBLE WRAPPER.

    This used to write to a separate api_execution_log.xlsx. That file is
    gone now - everything lives in the single consolidated Execution Report
    (REPORT_FILE) written by write_result(). This wrapper exists so any
    other code still calling write_api_log() directly (e.g.
    utils/api_logger.py, used by test_send_offer_validation.py for
    per-call logging in multi-request tests) keeps working without
    resurrecting a second file.

    NOTE: this writes its OWN row, separate from the one-row-per-test
    summary that the report_result fixture in conftest.py writes. If a
    test calls this once per API request (e.g. a validation test firing
    several negative-case requests), you'll correctly get one detailed
    row per request PLUS one summary row for the overall test result -
    that's intentional, not a duplicate-row bug.
    """

    write_result(
        test_name=test_name,
        status=result or "",
        action="",
        candidate_name="",
        candidate_email="",
        api_message="",
        run_id=run_id or "",
        environment=environment or "",
        username=username or "",
        method=method,
        endpoint=endpoint,
        api_status=(
            status_code
            if status_code is not None
            else (actual_status or "")
        ),
        expected_status=expected_status or "",
        duration=duration or "",
        sla=sla or "",
        sla_status=sla_status or "",
        request_headers=request_headers or "",
        request_payload=request_payload or "",
        response_body=response_body or "",
        error=error or "",
        screenshot="",
    )


# ==========================================================
# ONBOARDING DATA LOG
# ONE TEST EXECUTION = ONE ROW
# (Unchanged - separate concern from the execution report above.
# Not currently wired up in conftest.py; left as-is.)
# ==========================================================

def create_field_log():

    if not os.path.exists(FIELD_LOG_FILE):

        wb = Workbook()

        ws = wb.active

        ws.title = "Onboarding Data Log"

        ws.append([
            "Execution Time",
            "Test Name",

            # Employee Details
            "First Name",
            "Middle Name",
            "Last Name",
            "Email",

            # Job Details
            "Job Offered",
            "Department",
            "Job Title",
            "Manager",

            # Employment Details
            "Employment Type",
            "Company Entity",
            "Salary Structure",

            # Additional Information
            "Section"
        ])

        wb.save(FIELD_LOG_FILE)

        print(
            f"[EXCEL CREATED] {FIELD_LOG_FILE}"
        )


def write_field_log(
    test_name,
    field_data,
    section="Onboarding"
):

    create_field_log()

    wb = load_workbook(FIELD_LOG_FILE)
    ws = wb["Onboarding Data Log"]

    # ==========================================
    # Existing headers
    # ==========================================
    headers = [
        cell.value
        for cell in ws[1]
    ]

    # ==========================================
    # Required base columns
    # ==========================================
    base_data = {
    "execution_time": str(datetime.now()),
    "test_name": test_name,
    **field_data,
    "section": section
    }

    # ==========================================
    # Add new columns automatically
    # ==========================================
    for key in base_data.keys():

        header = key.replace("_", " ").title()

        if header not in headers:

            ws.cell(
                row=1,
                column=len(headers) + 1,
                value=header
            )

            headers.append(header)

    # ==========================================
    # Build row according to headers
    # ==========================================
    row = []

    for header in headers:

        lookup_key = (
            header
            .lower()
            .replace(" ", "_")
        )

        row.append(
            base_data.get(
                lookup_key,
                ""
            )
        )

    print("\n===== FIELD DATA RECEIVED =====")

    for k, v in field_data.items():
        print(k, "=", v)

    print("Field Count =", len(field_data))

    ws.append(row)

    print(
        f"\nWriting row to Excel. "
        f"Row Length={len(row)} "
        f"Headers={len(headers)}"
    )

    wb.save(FIELD_LOG_FILE)
    print(
        f"[FIELD LOG UPDATED] {FIELD_LOG_FILE}"
    )

def create_sql_injection_report():
    create_excel_report()

    wb = load_workbook(REPORT_FILE)

    if SQLI_SHEET_NAME in wb.sheetnames:
        wb.close()
        return

    ws = wb.create_sheet(SQLI_SHEET_NAME)

    headers = [
        "Execution Time",
        "Run ID",
        "Environment",
        "Username",
        "Test Name",
        "Field",
        "Payload Type",
        "Payload",
        "Candidate Name",
        "Candidate Email",
        "Method",
        "Endpoint",
        "Expected Status",
        "Actual Status",
        "Result",
        "Duration(ms)",
        "SLA(ms)",
        "SLA Status",
        "API Message",
        "Request Headers",
        "Request Payload",
        "Response Body",
        "Screenshot",
        "Error",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    ws.freeze_panes = "A2"

    wb.save(REPORT_FILE)

def _excel_safe(value):
    """Convert complex Python objects into Excel-compatible values."""
    if isinstance(value, (dict, list, tuple, set)):
        try:
            return json.dumps(value, ensure_ascii=False, default=str)
        except Exception:
            return str(value)
    return value


def write_sql_injection_result(
    field,
    payload,
    payload_type=None,
    execution_time=None,
    run_id=None,
    environment=None,
    username=None,
    test_name=None,
    candidate_name=None,
    candidate_email=None,
    method=None,
    endpoint=None,
    expected_status=None,
    actual_status=None,
    result=None,
    duration_ms=None,
    sla_ms=None,
    sla_status=None,
    api_message=None,
    request_headers=None,
    request_payload=None,
    response_body=None,
    screenshot=None,
    error=None,
):
    create_sql_injection_report()

    wb = load_workbook(REPORT_FILE)

    ws = wb[SQLI_SHEET_NAME]

    # -----------------------------------------------------------------------
    # EXCEL-SAFE VALUE CONVERSION
    if isinstance(payload, (dict, list)):
        try:
            payload = json.dumps(payload, ensure_ascii=False)
        except Exception:
            payload = str(payload)

        
    payload = _excel_safe(payload)
    request_headers = _excel_safe(request_headers)
    request_payload = _excel_safe(request_payload)
    response_body = _excel_safe(response_body)
    api_message = _excel_safe(api_message)
    error = _excel_safe(error)

    row = [
        execution_time or datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        ),
        run_id,
        environment,
        username,
        test_name,
        field,
        payload_type,
        payload,
        candidate_name,
        candidate_email,
        method,
        endpoint,
        expected_status,
        actual_status,
        result,
        duration_ms,
        sla_ms,
        sla_status,
        api_message,
        request_headers,
        request_payload,
        response_body,
        screenshot,
        error,
    ]

    ws.append(row)

    for cell in ws[ws.max_row]:
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )

    wb.save(REPORT_FILE)

    print(
        f"[SQLI REPORT UPDATED] {REPORT_FILE}"
    )