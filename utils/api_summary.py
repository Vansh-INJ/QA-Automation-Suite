from openpyxl import load_workbook
import os

from utils.run_manager import (
    get_run_folder
)

RUN_FOLDER = get_run_folder()

API_FILE = os.path.join(
    RUN_FOLDER,
    "excel",
    "api_execution_log.xlsx"
)


def print_api_summary():

    if not os.path.exists(
            API_FILE
    ):
        return

    wb = load_workbook(
        API_FILE
    )

    ws = wb[
        "API Execution Log"
    ]

    total = 0
    passed = 0
    failed = 0

    durations = []

    failed_tests = []

    for row in ws.iter_rows(
            min_row=2,
            values_only=True
    ):

        total += 1

        test_name = row[4]
        duration = row[11]
        result = row[13]
        expected = row[8]
        actual = row[9]

        if duration:
            durations.append(
                float(duration)
            )

        if result == "PASS":
            passed += 1

        elif result == "FAIL":

            failed += 1

            failed_tests.append({
                "test":
                    test_name,
                "expected":
                    expected,
                "actual":
                    actual
            })

    average = (
        round(
            sum(durations)
            / len(durations),
            2
        )
        if durations
        else 0
    )

    pass_percent = (
        round(
            passed
            / total
            * 100,
            2
        )
        if total
        else 0
    )

    print(
        "\n"
        + "=" * 55
    )

    print(
        "API EXECUTION SUMMARY"
    )

    print(
        "=" * 55
    )

    print(
        f"Run Folder : "
        f"{os.path.basename(RUN_FOLDER)}"
    )

    print(
        f"Total APIs : {total}"
    )

    print(
        f"Passed     : {passed}"
    )

    print(
        f"Failed     : {failed}"
    )

    print(
        f"Pass %     : "
        f"{pass_percent}%"
    )

    print(
        f"Average RT : "
        f"{average} ms"
    )

    if failed_tests:

        print(
            "\nFAILED TESTS"
        )

        print(
            "-" * 55
        )

        for item in failed_tests:

            print(
                f"\n{item['test']}"
            )

            print(
                f"Expected : "
                f"{item['expected']}"
            )

            print(
                f"Actual   : "
                f"{item['actual']}"
            )

    print(
        "\n"
        + "=" * 55
    )